"""
create_paper_tables.py
Generate Tables 4–9 for Section 5 of the CAB persistency paper.
All tables use cluster-robust standard errors (clustered at country level).
Output: paper_tables_section5.xlsx (in updated_2025/)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────────
BASE = "c:/Users/eclow/Documents/GitHub/cab-persistency/updated_2025/"
BASELINE_XL  = BASE + "mnlogit_baseline_clustered.xlsx"
MEMO_XL      = BASE + "memo_mnl_coeff_table_top_specs_with_yearfe.xlsx"
BINARY_XL    = BASE + "mnlogit_binary_robustness.xlsx"
SUBGROUP_XL  = BASE + "mnlogit_subgroups.xlsx"
OUT_XL       = BASE + "paper_tables_section5.xlsx"

# ── helpers ────────────────────────────────────────────────────────────────────
thin  = Side(style="thin")
thick = Side(style="medium")

def hdr_font(bold=True, sz=11):
    return Font(bold=bold, size=sz, name="Times New Roman")

def body_font(sz=11):
    return Font(size=sz, name="Times New Roman")

def border(top=None, bottom=None, left=None, right=None):
    return Border(top=top, bottom=bottom, left=left, right=right)

def top_thick():
    return Border(top=thick)

def bottom_thick():
    return Border(bottom=thick)

def top_thin():
    return Border(top=thin)

def bottom_thin():
    return Border(bottom=thin)

def write_cell(ws, row, col, value, bold=False, italic=False,
               align="center", top=None, bottom=None,
               left=None, right=None, wrap=False, sz=11, indent=0):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, italic=italic, size=sz, name="Times New Roman")
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap, indent=indent)
    c.border = Border(top=top, bottom=bottom, left=left, right=right)
    return c

def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(size=10, italic=True, name="Times New Roman")
    c.alignment = Alignment(horizontal="left", wrap_text=True)
    return c

# ══════════════════════════════════════════════════════════════════════════════
# TABLE 4 – Baseline MNL Coefficient Estimates
# ══════════════════════════════════════════════════════════════════════════════
def build_table4(wb):
    ws = wb.create_sheet("Table 4")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    r = 1
    # Title
    ws.merge_cells(f"A{r}:C{r}")
    write_cell(ws, r, 1,
               "Table 4. Multinomial Logit Estimates: Baseline Specification",
               bold=True, align="center", sz=12)
    r += 1

    ws.merge_cells(f"A{r}:C{r}")
    write_cell(ws, r, 1,
               "(Reference category: Explosive; cluster-robust standard errors in parentheses)",
               italic=True, align="center", sz=10)
    r += 1
    r += 1  # blank

    # Column headers
    write_cell(ws, r, 1, "Variable",  bold=True, align="left",
               bottom=thin, top=thick, sz=11)
    write_cell(ws, r, 2, "Stationary", bold=True, align="center",
               bottom=thin, top=thick, sz=11)
    write_cell(ws, r, 3, "Unit root",  bold=True, align="center",
               bottom=thin, top=thick, sz=11)
    r += 1

    # Data
    rows = [
        ("Constant",                     "3.085**",  "3.554**"),
        ("",                             "(1.087)",  "(1.148)"),
        ("Fixed exchange rate",          "1.869***", "1.980***"),
        ("",                             "(0.556)",  "(0.549)"),
        ("Current account deficit",      "0.842*",   "0.728"),
        ("",                             "(0.481)",  "(0.503)"),
        ("Commodity exporter",           "0.188",    "−0.137"),
        ("",                             "(0.759)",  "(0.757)"),
        ("Manufacturing exporter",       "−1.204*",  "−1.477**"),
        ("",                             "(0.683)",  "(0.670)"),
        ("Trade openness (t−1)",         "−1.273**", "−0.819*"),
        ("",                             "(0.557)",  "(0.453)"),
        ("Fiscal balance (3-yr avg, t−1)", "1.657",  "3.241*"),
        ("",                             "(1.933)",  "(1.809)"),
    ]

    for label, stat, ur in rows:
        is_se = label == ""
        for col, val in zip([1, 2, 3], [label, stat, ur]):
            write_cell(ws, r, col, val,
                       italic=is_se,
                       align="left" if col == 1 else "center",
                       sz=11,
                       indent=1 if (is_se and col == 1) else 0)
        r += 1

    # Stats rows
    write_cell(ws, r, 1, "", top=thin)
    write_cell(ws, r, 2, "", top=thin)
    write_cell(ws, r, 3, "", top=thin)
    r += 1

    stats = [
        ("Observations",    "2,451", "2,451"),
        ("Countries",       "74",    "74"),
        ("McFadden R²",     "0.050", "0.050"),
        ("BIC per obs.",    "1.661", "1.661"),
        ("Log-likelihood",  "−1,980.3", "−1,980.3"),
    ]
    for i, (label, stat, ur) in enumerate(stats):
        is_last = i == len(stats) - 1
        bot = thick if is_last else None
        write_cell(ws, r, 1, label, align="left", bottom=bot, sz=11)
        write_cell(ws, r, 2, stat,  align="center", bottom=bot, sz=11)
        write_cell(ws, r, 3, ur,    align="center", bottom=bot, sz=11)
        r += 1

    r += 1
    note(ws, r, 1,
         "Notes: *p < 0.10, **p < 0.05, ***p < 0.01. Standard errors clustered at the country level. "
         "All covariates are lagged or pre-determined at t−1. Continuous regressors are standardised "
         "(mean zero, unit standard deviation) at the sample level. "
         "The explosive regime serves as the reference category.")


# ══════════════════════════════════════════════════════════════════════════════
# TABLE 5 – Average Marginal Effects
# ══════════════════════════════════════════════════════════════════════════════
def build_table5(wb):
    ws = wb.create_sheet("Table 5")
    ws.column_dimensions["A"].width = 34
    for col in ["B","C","D"]:
        ws.column_dimensions[col].width = 18

    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    write_cell(ws, r, 1,
               "Table 5. Average Marginal Effects: Baseline Specification",
               bold=True, align="center", sz=12)
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    write_cell(ws, r, 1,
               "(Standard errors in parentheses; computed at sample means, cluster-robust)",
               italic=True, align="center", sz=10)
    r += 1
    r += 1

    # Headers
    hdrs = ["Variable", "ΔP(Explosive)", "ΔP(Stationary)", "ΔP(Unit root)"]
    for col, h in enumerate(hdrs, 1):
        write_cell(ws, r, col, h, bold=True,
                   align="left" if col==1 else "center",
                   bottom=thin, top=thick)
    r += 1

    # Data: dy_dx and SE per variable
    data = [
        # (label, exp_dy, exp_se, exp_stars, stat_dy, stat_se, stat_stars, ur_dy, ur_se, ur_stars)
        ("Fixed exchange rate",
         "-0.091", "(0.045)", "**",
         "0.011",  "(0.071)", "",
         "0.080",  "(0.076)", ""),
        ("Current account deficit",
         "-0.036", "(0.026)", "",
         "0.040",  "(0.057)", "",
         "-0.003", "(0.063)", ""),
        ("Commodity exporter",
         "0.000",  "(0.034)", "",
         "0.072",  "(0.098)", "",
         "-0.073", "(0.102)", ""),
        ("Manufacturing exporter",
         "0.064",  "(0.040)", "",
         "0.035",  "(0.095)", "",
         "-0.100", "(0.098)", ""),
        ("Trade openness (t−1)",
         "0.047",  "(0.019)", "**",
         "-0.120", "(0.112)", "",
         "0.073",  "(0.110)", ""),
        ("Fiscal balance (3-yr avg, t−1)",
         "-0.123", "(0.079)", "",
         "-0.305", "(0.318)", "",
         "0.428",  "(0.318)", ""),
    ]

    for label, exp_dy, exp_se, exp_st, stat_dy, stat_se, stat_st, ur_dy, ur_se, ur_st in data:
        # Coefficient row
        write_cell(ws, r, 1, label,       align="left")
        write_cell(ws, r, 2, exp_dy+exp_st,   align="center")
        write_cell(ws, r, 3, stat_dy+stat_st, align="center")
        write_cell(ws, r, 4, ur_dy+ur_st,    align="center")
        r += 1
        # SE row
        write_cell(ws, r, 1, "",          align="left", indent=1)
        write_cell(ws, r, 2, exp_se,  align="center", italic=True)
        write_cell(ws, r, 3, stat_se, align="center", italic=True)
        write_cell(ws, r, 4, ur_se,   align="center", italic=True)
        r += 1

    # Separator + stats
    for col in [1,2,3,4]:
        write_cell(ws, r, col, "", top=thin)
    r += 1
    write_cell(ws, r, 1, "Observations", align="left")
    for col, v in zip([2,3,4], ["2,451","2,451","2,451"]):
        write_cell(ws, r, col, v, align="center")
    r += 1
    write_cell(ws, r, 1, "Countries", align="left")
    for col in [2,3,4]:
        write_cell(ws, r, col, "74", align="center")
    r += 1
    for col in [1,2,3,4]:
        write_cell(ws, r, col, "", bottom=thick)
    r += 1
    note(ws, r, 1,
         "Notes: *p < 0.10, **p < 0.05, ***p < 0.01. Average marginal effects computed "
         "as the sample mean of individual marginal effects (at='overall'). "
         "Standard errors clustered at the country level. "
         "Entries are changes in predicted probabilities from a unit change in the regressor.")


# ══════════════════════════════════════════════════════════════════════════════
# TABLE 6 – Model Specification Comparison
# ══════════════════════════════════════════════════════════════════════════════
def build_table6(wb):
    ws = wb.create_sheet("Table 6")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10

    r = 1
    ws.merge_cells(f"A{r}:F{r}")
    write_cell(ws, r, 1,
               "Table 6. Model Specification Comparison",
               bold=True, align="center", sz=12)
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    write_cell(ws, r, 1,
               "(Cluster-robust standard errors throughout; baseline model in bold)",
               italic=True, align="center", sz=10)
    r += 1
    r += 1

    hdrs = ["Specification", "Key features", "N", "McFadden R²", "BIC/obs", "Year FE"]
    for col, h in enumerate(hdrs, 1):
        write_cell(ws, r, col, h, bold=True,
                   align="left" if col <= 2 else "center",
                   bottom=thin, top=thick)
    r += 1

    rows = [
        ("C00004",            "Contemporaneous trade openness",                         "2,613", "0.048", "1.663", "No"),
        ("C00004 + YFE",      "Full year dummies (failed to converge)",                 "2,613", "0.058", "1.958", "Yes"),
        ("C00058 (baseline)", "Lagged openness + fiscal balance",                       "2,451", "0.050", "1.661", "No"),
        ("C00058 + YFE",      "Full year dummies",                                      "2,451", "0.063", "1.970", "Yes"),
        ("C00342_fdgap",      "+ Financial development gap",                            "2,064", "0.067", "1.670", "No"),
        ("C00058_crisis",     "+ GFC (2008–09) and COVID (2020–21) dummies",            "2,451", "0.051", "1.673", "No"),
    ]

    baseline_row = 2  # 0-indexed from rows list

    for i, (spec, feat, n, r2, bic, yfe) in enumerate(rows):
        is_base = i == baseline_row
        is_last = i == len(rows) - 1
        bot = thick if is_last else None

        def wc(col, val):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(bold=is_base, size=11, name="Times New Roman")
            c.alignment = Alignment(
                horizontal="left" if col <= 2 else "center",
                vertical="center"
            )
            if is_last:
                c.border = Border(bottom=thick)
            return c

        wc(1, spec); wc(2, feat); wc(3, n); wc(4, r2); wc(5, bic); wc(6, yfe)
        r += 1

    r += 1
    note(ws, r, 1,
         "Notes: C00058 is the preferred specification. YFE = year fixed effects. "
         "C00058_oil (oil exporter indicator) cannot be estimated: zero oil-exporter observations "
         "fall in the unit root regime (complete separation), leaving the unit root equation unidentified. "
         "The oil cross-tabulation (13 explosive / 27 stationary / 0 unit root) is reported descriptively. "
         "BIC/obs = Bayesian Information Criterion divided by sample size. "
         "Lower BIC/obs indicates better fit, penalised for model complexity.")


# ══════════════════════════════════════════════════════════════════════════════
# TABLE 7 – Binary Logit Robustness
# ══════════════════════════════════════════════════════════════════════════════
def build_table7(wb):
    ws = wb.create_sheet("Table 7")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12

    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    write_cell(ws, r, 1,
               "Table 7. Binary Logit Robustness: Stationary vs. Unit Root",
               bold=True, align="center", sz=12)
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    write_cell(ws, r, 1,
               "(134 explosive observations dropped; outcome: 1 = Unit root, 0 = Stationary; "
               "average marginal effects; cluster-robust SEs)",
               italic=True, align="center", sz=10)
    r += 1
    r += 1

    hdrs = ["Variable", "AME", "Std. Error", "p-value"]
    for col, h in enumerate(hdrs, 1):
        write_cell(ws, r, col, h, bold=True,
                   align="left" if col==1 else "center",
                   bottom=thin, top=thick)
    r += 1

    data = [
        ("Fixed exchange rate",            "0.027",  "0.075", "0.717"),
        ("Current account deficit",       "-0.027",  "0.062", "0.669"),
        ("Commodity exporter",            "-0.075",  "0.102", "0.460"),
        ("Manufacturing exporter",        "-0.062",  "0.099", "0.531"),
        ("Trade openness (t−1)",           "0.100",  "0.113", "0.377"),
        ("Fiscal balance (3-yr avg, t−1)", "0.358",  "0.322", "0.267"),
    ]

    for i, (label, ame, se, pval) in enumerate(data):
        is_last = i == len(data) - 1
        bot = thick if is_last else None
        write_cell(ws, r, 1, label,  align="left", bottom=bot)
        write_cell(ws, r, 2, ame,    align="center", bottom=bot)
        write_cell(ws, r, 3, se,     align="center", bottom=bot)
        write_cell(ws, r, 4, pval,   align="center", bottom=bot)
        r += 1

    r += 1
    write_cell(ws, r, 1, "Observations",   align="left"); write_cell(ws, r, 2, "2,317", align="center"); r += 1
    write_cell(ws, r, 1, "Countries",      align="left"); write_cell(ws, r, 2, "74",    align="center"); r += 1

    r += 1
    note(ws, r, 1,
         "Notes: No variable is statistically significant (all p > 0.25). "
         "This result confirms that the primary explanatory power of the baseline MNL operates "
         "through the explosive-versus-non-explosive distinction. The stationary–unit-root split "
         "within the non-explosive sample is not explained by the same structural covariates.")


# ══════════════════════════════════════════════════════════════════════════════
# TABLE 8 – Income Subgroup Results
# ══════════════════════════════════════════════════════════════════════════════
def build_table8(wb):
    ws = wb.create_sheet("Table 8")
    ws.column_dimensions["A"].width = 34
    for col in ["B","C","D"]:
        ws.column_dimensions[col].width = 18

    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    write_cell(ws, r, 1,
               "Table 8. MNL Results by Income Group",
               bold=True, align="center", sz=12)
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    write_cell(ws, r, 1,
               "(Cluster-robust standard errors; baseline specification C00058; "
               "reference category: Explosive)",
               italic=True, align="center", sz=10)
    r += 1
    r += 1

    # Panel A: fit stats
    ws.merge_cells(f"A{r}:D{r}")
    write_cell(ws, r, 1, "Panel A: Sample and Fit Statistics",
               bold=True, align="left", bottom=thin, top=thick)
    r += 1

    hdrs = ["", "Advanced", "Emerging", "Low income"]
    for col, h in enumerate(hdrs, 1):
        write_cell(ws, r, col, h, bold=True,
                   align="left" if col==1 else "center", bottom=thin)
    r += 1

    fit_rows = [
        ("Observations",      "1,041",  "903",   "507"),
        ("Countries",         "23",     "28",    "23"),
        ("Explosive obs.",    "15",     "116",   "3"),
        ("Stationary obs.",   "472",    "299",   "172"),
        ("Unit root obs.",    "554",    "488",   "332"),
        ("McFadden R²",       "—",      "0.186", "—"),
        ("BIC per obs.",      "—",      "1.671", "—"),
        ("Model converged",   "No*",    "Yes",   "No*"),
    ]
    for i, row in enumerate(fit_rows):
        is_last = i == len(fit_rows) - 1
        bot = thin if is_last else None
        for col, val in enumerate(row, 1):
            write_cell(ws, r, col, val,
                       align="left" if col==1 else "center", bottom=bot)
        r += 1

    r += 1

    # Panel B: Emerging market coefficients
    ws.merge_cells(f"A{r}:D{r}")
    write_cell(ws, r, 1,
               "Panel B: Coefficient Estimates — Emerging Market Economies",
               bold=True, align="left", bottom=thin, top=thick)
    r += 1

    ws.merge_cells(f"A{r}:B{r}")
    write_cell(ws, r, 1, "Variable", bold=True, align="left", bottom=thin)
    write_cell(ws, r, 3, "Stationary", bold=True, align="center", bottom=thin)
    write_cell(ws, r, 4, "Unit root",  bold=True, align="center", bottom=thin)
    r += 1

    em_data = [
        ("Constant",                     "1.826",    "2.904**"),
        ("",                             "(1.210)",  "(1.330)"),
        ("Fixed exchange rate",          "0.815",    "1.259**"),
        ("",                             "(0.723)",  "(0.593)"),
        ("Current account deficit",      "1.338**",  "0.571"),
        ("",                             "(0.513)",  "(0.552)"),
        ("Commodity exporter",           "0.681",    "0.083"),
        ("",                             "(0.889)",  "(0.943)"),
        ("Manufacturing exporter",       "−0.392",   "−1.614**"),
        ("",                             "(0.752)",  "(0.758)"),
        ("Trade openness (t−1)",         "−2.641**", "−0.467"),
        ("",                             "(1.092)",  "(0.511)"),
        ("Fiscal balance (3-yr avg, t−1)", "−1.434", "5.778**"),
        ("",                             "(3.700)",  "(2.816)"),
    ]

    for label, stat, ur in em_data:
        is_se = label == ""
        ws.merge_cells(f"A{r}:B{r}")
        write_cell(ws, r, 1, label,
                   italic=is_se, align="left",
                   indent=1 if is_se else 0)
        write_cell(ws, r, 3, stat, italic=is_se, align="center")
        write_cell(ws, r, 4, ur,   italic=is_se, align="center")
        r += 1

    for col in [1,3,4]:
        write_cell(ws, r, col, "", top=thin, bottom=thick)
    r += 1
    write_cell(ws, r, 1, "Observations", align="left");   write_cell(ws, r, 3, "903", align="center"); r += 1
    write_cell(ws, r, 1, "Countries",    align="left");   write_cell(ws, r, 3, "28",  align="center"); r += 1
    write_cell(ws, r, 1, "McFadden R²",  align="left");   write_cell(ws, r, 3, "0.186", align="center"); r += 1
    for col in [1,3,4]:
        write_cell(ws, r, col, "", bottom=thick)
    r += 1

    r += 1
    note(ws, r, 1,
         "Notes: *p < 0.10, **p < 0.05, ***p < 0.01. Standard errors in parentheses, clustered at country level. "
         "* Advanced and low-income subgroup MNL estimates are not reported: the explosive regime contains "
         "only 15 and 3 observations respectively, insufficient to identify the explosive equation. "
         "Subgroup models are otherwise estimated on the same specification as the baseline (C00058).")


# ══════════════════════════════════════════════════════════════════════════════
# TABLE 9 – Time Period Subgroup Results
# ══════════════════════════════════════════════════════════════════════════════
def build_table9(wb):
    ws = wb.create_sheet("Table 9")
    ws.column_dimensions["A"].width = 34
    for col in ["B","C","D","E"]:
        ws.column_dimensions[col].width = 16

    r = 1
    ws.merge_cells(f"A{r}:E{r}")
    write_cell(ws, r, 1,
               "Table 9. MNL Results by Time Period",
               bold=True, align="center", sz=12)
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    write_cell(ws, r, 1,
               "(Cluster-robust standard errors; baseline specification C00058; "
               "reference category: Explosive)",
               italic=True, align="center", sz=10)
    r += 1
    r += 1

    # Panel A: fit stats
    ws.merge_cells(f"A{r}:E{r}")
    write_cell(ws, r, 1, "Panel A: Sample and Fit Statistics",
               bold=True, align="left", bottom=thin, top=thick)
    r += 1

    hdrs = ["", "Pre-GFC (< 2008)", "", "Post-GFC (≥ 2008)", ""]
    for col, h in enumerate(hdrs, 1):
        write_cell(ws, r, col, h, bold=True, align="center", bottom=thin)
    write_cell(ws, r, 1, "", bold=True, align="left", bottom=thin)
    r += 1

    fit_rows = [
        ("Observations",      "1,356",  "—",  "1,095",  "—"),
        ("Countries",         "70",     "—",  "74",     "—"),
        ("Explosive obs.",    "65",     "—",  "69",     "—"),
        ("McFadden R²",       "0.042",  "—",  "0.089",  "—"),
        ("BIC per obs.",      "1.692",  "—",  "1.645",  "—"),
    ]
    for i, row in enumerate(fit_rows):
        is_last = i == len(fit_rows) - 1
        bot = thin if is_last else None
        for col, val in enumerate(row, 1):
            write_cell(ws, r, col, val,
                       align="left" if col==1 else "center", bottom=bot)
        r += 1

    r += 1

    # Panel B: Coefficient table
    ws.merge_cells(f"A{r}:E{r}")
    write_cell(ws, r, 1, "Panel B: Coefficient Estimates",
               bold=True, align="left", bottom=thin, top=thick)
    r += 1

    hdrs2 = ["Variable",
             "Stat. (Pre-GFC)", "UR (Pre-GFC)",
             "Stat. (Post-GFC)", "UR (Post-GFC)"]
    for col, h in enumerate(hdrs2, 1):
        write_cell(ws, r, col, h, bold=True,
                   align="left" if col==1 else "center", bottom=thin)
    r += 1

    coef_data = [
        ("Constant",
         "3.460***", "4.065***", "3.723***", "3.954***"),
        ("",
         "(1.027)",  "(1.114)",  "(1.105)",  "(1.046)"),
        ("Fixed exchange rate",
         "0.979",    "1.206*",   "3.813***", "3.700***"),
        ("",
         "(0.689)",  "(0.698)",  "(0.817)",  "(0.775)"),
        ("Current account deficit",
         "0.020",    "0.018",    "2.076**",  "1.753*"),
        ("",
         "(0.557)",  "(0.566)",  "(0.924)",  "(0.924)"),
        ("Commodity exporter",
         "0.717",    "−0.100",   "−1.081",   "−0.745"),
        ("",
         "(0.810)",  "(0.795)",  "(0.680)",  "(0.653)"),
        ("Manufacturing exporter",
         "−1.199",   "−1.749*",  "−2.041**", "−1.894**"),
        ("",
         "(0.884)",  "(0.903)",  "(0.709)",  "(0.623)"),
        ("Trade openness (t−1)",
         "−1.282*",  "−0.774",   "−1.493**", "−1.032**"),
        ("",
         "(0.695)",  "(0.629)",  "(0.544)",  "(0.347)"),
        ("Fiscal balance (3-yr avg, t−1)",
         "−0.920",   "0.625",    "6.421**",  "7.632***"),
        ("",
         "(2.388)",  "(2.381)",  "(2.645)",  "(2.308)"),
    ]

    for label, s1, ur1, s2, ur2 in coef_data:
        is_se = label == ""
        write_cell(ws, r, 1, label, italic=is_se, align="left",
                   indent=1 if is_se else 0)
        write_cell(ws, r, 2, s1,  italic=is_se, align="center")
        write_cell(ws, r, 3, ur1, italic=is_se, align="center")
        write_cell(ws, r, 4, s2,  italic=is_se, align="center")
        write_cell(ws, r, 5, ur2, italic=is_se, align="center")
        r += 1

    for col in range(1, 6):
        write_cell(ws, r, col, "", top=thin, bottom=thick)
    r += 1
    for label, v1, v2 in [
        ("Observations",  "1,356", "1,095"),
        ("Countries",     "70",    "74"),
        ("McFadden R²",   "0.042", "0.089"),
        ("BIC per obs.",  "1.692", "1.645"),
    ]:
        write_cell(ws, r, 1, label, align="left")
        write_cell(ws, r, 2, v1, align="center")
        write_cell(ws, r, 3, v1, align="center")
        write_cell(ws, r, 4, v2, align="center")
        write_cell(ws, r, 5, v2, align="center")
        r += 1

    for col in range(1, 6):
        write_cell(ws, r, col, "", bottom=thick)
    r += 1

    r += 1
    note(ws, r, 1,
         "Notes: *p < 0.10, **p < 0.05, ***p < 0.01. Standard errors in parentheses, clustered at country level. "
         "Pre-GFC = 1985–2007; Post-GFC = 2008–2023. "
         "The post-GFC subsample includes the COVID-19 recession (2020–21). "
         "The increase in McFadden R² post-GFC (0.042 → 0.089) and the stronger coefficients on "
         "fixed exchange rate and fiscal balance suggest that structural determinants have become "
         "more potent predictors of regime classification since the Global Financial Crisis.")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    wb = Workbook()
    wb.remove(wb.active)          # remove default blank sheet

    build_table4(wb)
    build_table5(wb)
    build_table6(wb)
    build_table7(wb)
    build_table8(wb)
    build_table9(wb)

    wb.save(OUT_XL)
    print(f"Saved: {OUT_XL}")
    print("Sheets:", [ws.title for ws in wb.worksheets])
