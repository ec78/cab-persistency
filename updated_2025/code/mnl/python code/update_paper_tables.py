"""
update_paper_tables.py
Updates paper_tables_section5.xlsx with:
  - Rebuilt Table 6 (adds lagged-ER row + CV log-loss column)
  - New Table 4a (lagged ER coefficient table)
  - New Table C1 (confusion matrix + predictive accuracy)
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

BASE   = "c:/Users/eclow/Documents/GitHub/cab-persistency/updated_2025/"
OUT_XL = BASE + "paper_tables_section5.xlsx"

thick = Side(style="medium")
thin  = Side(style="thin")


def wc(ws, row, col, value, bold=False, italic=False,
        align="center", top=None, bottom=None, sz=11, indent=0, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, italic=italic, size=sz, name="Times New Roman")
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap, indent=indent)
    c.border    = Border(top=top, bottom=bottom)
    return c


def note_cell(ws, row, col, text, ncols=4):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row,   end_column=col + ncols - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(size=10, italic=True, name="Times New Roman")
    c.alignment = Alignment(horizontal="left", wrap_text=True)
    return c


# ===========================================================================
# TABLE 6  (rebuild)
# ===========================================================================
def build_table6(wb):
    if "Table 6" in wb.sheetnames:
        del wb["Table 6"]
    ws = wb.create_sheet("Table 6")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    r = 1
    ws.merge_cells(f"A{r}:G{r}")
    wc(ws, r, 1, "Table 6. Model Specification Comparison",
       bold=True, align="center", sz=12)
    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    wc(ws, r, 1,
       "(Cluster-robust SEs throughout; baseline in bold; "
       "CV log-loss = 5-fold country-grouped cross-validation)",
       italic=True, align="center", sz=10)
    r += 2

    for col, h in enumerate(
        ["Specification", "Key features", "N", "McFadden R2", "BIC/obs",
         "Year FE", "CV log-loss"], 1
    ):
        wc(ws, r, col, h, bold=True,
           align="left" if col <= 2 else "center",
           bottom=thin, top=thick)
    r += 1

    rows = [
        ("C00004",            "Contemporaneous trade openness",               "2,613", "0.048", "1.663", "No",  "—",    False),
        ("C00004 + YFE",      "Full year dummies (convergence failure)",       "2,613", "0.058", "1.958", "Yes", "—",    False),
        ("C00058 (baseline)", "Lagged openness + fiscal balance",              "2,451", "0.050", "1.661", "No",  "0.945",True),
        ("C00058 + YFE",      "Full year dummies",                             "2,451", "0.063", "1.970", "Yes", "—",    False),
        ("C00058_er_lag",     "Lagged ER dummy (endogeneity check)",           "2,451", "0.051", "1.659", "No",  "0.945",False),
        ("C00342_fdgap",      "+ Financial development gap",                   "2,064", "0.067", "1.670", "No",  "—",    False),
        ("C00058_crisis",     "+ GFC (2008-09) and COVID (2020-21) dummies",   "2,451", "0.051", "1.673", "No",  "—",    False),
    ]

    for i, (spec, feat, n, r2, bic, yfe, cvll, is_base) in enumerate(rows):
        is_last = (i == len(rows) - 1)
        for col, val in enumerate([spec, feat, n, r2, bic, yfe, cvll], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(bold=is_base, size=11, name="Times New Roman")
            c.alignment = Alignment(
                horizontal="left" if col <= 2 else "center",
                vertical="center"
            )
            if is_last:
                c.border = Border(bottom=thick)
        r += 1

    r += 1
    note_cell(ws, r, 1,
        "Notes: BIC/obs = BIC / n (lower = better). "
        "CV log-loss = mean log-loss, country-grouped 5-fold CV (lower = better). "
        "C00058_er_lag replaces contemporaneous fixed_er with its one-year within-country lag; "
        "results are virtually identical to C00058 (fixed_er AME = -0.094** vs -0.091**), "
        "confirming robustness to reverse causality. "
        "C00058_oil cannot be estimated (complete separation in unit root equation). "
        "Year FE = annual time dummies. Dashes = not estimated or not applicable.",
        ncols=7)


# ===========================================================================
# TABLE 4a  (new: lagged ER coefficients)
# ===========================================================================
def build_table4a(wb):
    if "Table 4a" in wb.sheetnames:
        del wb["Table 4a"]
    ws = wb.create_sheet("Table 4a")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    r = 1
    ws.merge_cells(f"A{r}:C{r}")
    wc(ws, r, 1,
       "Table 4a. Robustness: MNL Estimates with Lagged Exchange Rate Dummy",
       bold=True, align="center", sz=12)
    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    wc(ws, r, 1,
       "(Specification C00058_er_lag; fixed_er lagged one year; "
       "cluster-robust SEs in parentheses; reference = Explosive)",
       italic=True, align="center", sz=10)
    r += 2

    wc(ws, r, 1, "Variable",   bold=True, align="left",   top=thick, bottom=thin)
    wc(ws, r, 2, "Stationary", bold=True, align="center", top=thick, bottom=thin)
    wc(ws, r, 3, "Unit root",  bold=True, align="center", top=thick, bottom=thin)
    r += 1

    coef_rows = [
        ("Constant",                       "3.420**",  "2.958**"),
        ("",                               "(1.152)",  "(1.090)"),
        ("Fixed exchange rate (t-1)",      "2.039***", "1.965***"),
        ("",                               "(0.487)",  "(0.491)"),
        ("Current account deficit",        "0.715",    "0.830*"),
        ("",                               "(0.515)",  "(0.494)"),
        ("Commodity exporter",             "-0.026",   "0.288"),
        ("",                               "(0.775)",  "(0.776)"),
        ("Manufacturing exporter",         "-1.328**", "-1.071"),
        ("",                               "(0.656)",  "(0.665)"),
        ("Trade openness (t-1)",           "-0.824*",  "-1.290**"),
        ("",                               "(0.462)",  "(0.565)"),
        ("Fiscal balance (3-yr avg, t-1)", "3.281*",   "1.699"),
        ("",                               "(1.882)",  "(2.002)"),
    ]

    for label, stat, ur in coef_rows:
        is_se = (label == "")
        for col, val in zip([1, 2, 3], [label, stat, ur]):
            wc(ws, r, col, val, italic=is_se,
               align="left" if col == 1 else "center",
               indent=1 if (is_se and col == 1) else 0)
        r += 1

    for col in [1, 2, 3]:
        wc(ws, r, col, "", top=thin)
    r += 1

    stats = [
        ("Observations",        "2,451"),
        ("Countries",           "74"),
        ("McFadden R2",         "0.051"),
        ("BIC per obs.",        "1.659"),
        ("CV log-loss (5-fold)","0.945"),
    ]
    for i, (lbl, val) in enumerate(stats):
        is_last = (i == len(stats) - 1)
        bot = thick if is_last else None
        wc(ws, r, 1, lbl, align="left",   bottom=bot)
        wc(ws, r, 2, val, align="center", bottom=bot)
        wc(ws, r, 3, val, align="center", bottom=bot)
        r += 1

    r += 1
    note_cell(ws, r, 1,
        "Notes: *p < 0.10, **p < 0.05, ***p < 0.01. Cluster-robust SEs in parentheses. "
        "Coefficients are essentially unchanged vs Table 4 (contemporaneous fixed_er). "
        "The fixed exchange rate effect is marginally larger and more precisely estimated "
        "when lagged (beta = 1.97-2.04 vs 1.87-1.98), confirming the result is not an "
        "artifact of contemporaneous reverse causality.",
        ncols=3)


# ===========================================================================
# TABLE C1  (new: confusion matrix + predictive accuracy)
# ===========================================================================
def build_table_c1(wb):
    if "Table C1" in wb.sheetnames:
        del wb["Table C1"]
    ws = wb.create_sheet("Table C1")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 32

    r = 1
    ws.merge_cells(f"A{r}:F{r}")
    wc(ws, r, 1,
       "Table C1. Predictive Performance: Baseline Specification (C00058)",
       bold=True, align="center", sz=12)
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    wc(ws, r, 1,
       "(In-sample predictions, modal classification rule; "
       "5-fold country-grouped cross-validation)",
       italic=True, align="center", sz=10)
    r += 2

    # Panel A: Confusion Matrix
    ws.merge_cells(f"A{r}:F{r}")
    wc(ws, r, 1, "Panel A: Confusion Matrix (Rows = Observed; Columns = Predicted)",
       bold=True, align="left", top=thick, bottom=thin)
    r += 1

    for col, h in enumerate(
        ["", "Pred: Explosive", "Pred: Stationary", "Pred: Unit root", "Total"], 1
    ):
        wc(ws, r, col, h, bold=True,
           align="left" if col == 1 else "center", bottom=thin)
    r += 1

    cm = [
        ("Obs: Explosive",   "4",   "0",    "130",  "134"),
        ("Obs: Stationary",  "0",   "164",  "779",  "943"),
        ("Obs: Unit root",   "8",   "68",   "1,298","1,374"),
        ("Total (predicted)","12",  "232",  "2,207","2,451"),
    ]
    for i, row_data in enumerate(cm):
        is_sep = (i == len(cm) - 2)
        is_last= (i == len(cm) - 1)
        bot = (thick if is_last else (thin if is_sep else None))
        for col, val in enumerate(row_data, 1):
            wc(ws, r, col, val,
               bold=is_last,
               align="left" if col == 1 else "center",
               bottom=bot)
        r += 1

    r += 1

    # Panel B: Summary accuracy
    ws.merge_cells(f"A{r}:F{r}")
    wc(ws, r, 1, "Panel B: Accuracy Statistics",
       bold=True, align="left", top=thick, bottom=thin)
    r += 1

    for col, h in enumerate(["Metric", "Value", "", "", "", ""], 1):
        wc(ws, r, col, h, bold=True,
           align="left" if col == 1 else "center", bottom=thin)
    r += 1

    acc = [
        ("Overall in-sample accuracy", "59.8%"),
        ("Base-rate accuracy (always predict modal)",   "56.1%"),
        ("Accuracy gain over base rate",               "+3.7 pp"),
        ("Mean CV log-loss (5-fold, country-grouped)", "0.945"),
    ]
    for label, val in acc:
        ws.merge_cells(f"A{r}:A{r}")
        wc(ws, r, 1, label, align="left")
        wc(ws, r, 2, val,   align="center")
        r += 1

    ws.merge_cells(f"A{r}:F{r}")
    wc(ws, r, 1, "", bottom=thin)
    r += 1

    # Panel C: Per-class stats
    ws.merge_cells(f"A{r}:F{r}")
    wc(ws, r, 1, "Panel C: Per-Class Precision and Recall",
       bold=True, align="left", bottom=thin)
    r += 1

    for col, h in enumerate(
        ["Outcome", "N (actual)", "Precision", "Recall", "F1", "Interpretation"], 1
    ):
        wc(ws, r, col, h, bold=True,
           align="left" if col in [1, 6] else "center", bottom=thin)
    r += 1

    pc = [
        ("Explosive",  "134",  "33%", "3%",  "0.05",
         "Model rarely predicts explosive (imbalanced, 5.5% base rate)"),
        ("Stationary", "943",  "71%", "17%", "0.27", ""),
        ("Unit root",  "1,374","59%", "95%", "0.72",
         "Dominates predictions; modal category (56.1% share)"),
    ]
    for i, (cat, n, prec, rec, f1, interp) in enumerate(pc):
        is_last = (i == len(pc) - 1)
        bot = thick if is_last else None
        for col, val in enumerate([cat, n, prec, rec, f1, interp], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(size=11, name="Times New Roman")
            c.alignment = Alignment(
                horizontal="left" if col in [1, 6] else "center",
                vertical="center", wrap_text=(col == 6)
            )
            if is_last:
                c.border = Border(bottom=thick)
        r += 1

    r += 1
    note_cell(ws, r, 1,
        "Notes: Predictions use the modal classification rule (argmax of predicted probabilities). "
        "Low explosive recall is expected for an MNL with a 5.5% explosive base rate; the model "
        "captures the conditional probability gradient rather than functioning as a balanced "
        "classifier. CV log-loss = mean log-loss, country-grouped 5-fold cross-validation "
        "(baseline 0.945 vs lagged-ER 0.945 -- virtually identical). "
        "F1 = 2 x (precision x recall) / (precision + recall).",
        ncols=6)


# ===========================================================================
# MAIN
# ===========================================================================
if __name__ == "__main__":
    wb = load_workbook(OUT_XL)

    build_table6(wb)
    build_table4a(wb)
    build_table_c1(wb)

    # Reorder sheets
    desired = ["Table 4", "Table 4a", "Table 5", "Table 6",
               "Table 7", "Table 8", "Table 9", "Table C1"]
    # openpyxl: move each sheet to its target position
    for target_pos, name in enumerate(desired):
        if name in wb.sheetnames:
            current_pos = wb.sheetnames.index(name)
            wb.move_sheet(wb[name], offset=target_pos - current_pos)

    wb.save(OUT_XL)
    print(f"Saved: {OUT_XL}")
    print("Sheets:", wb.sheetnames)
